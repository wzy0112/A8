import os
import sys
import csv
import math
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

# feq=1h
SUMO_HOME = r"D:\Eclipse\SUMO"

# 之前只有5个 而且是某一个lane上的 但是BASt网站里面每个detector给了2个richtunged 所以我必须给每个detector匹配上相邻反方向道路也放一个 一共10个
# 之前10个detecotrs只占据单独lane， 但是但是BASt网站里面每个detector给了2个richtungen, 对应的是整个edge的数据

# 一行表格 = 一个测点组
# Richtung1_TK-Blatt + Richtung1_BASt-Nr. = sensor_id 基础编号
# Richtung1_lon/lat -> dir1
# Richtung2_lon/lat -> dir2
# 最终 detector:
# 77349810_dir1_lane0
# 77349810_dir2_lane0

def load_sumo_tools():
    tools_path = os.path.join(SUMO_HOME, "tools")
    if tools_path not in sys.path:
        sys.path.append(tools_path)

    try:
        import sumolib  # noqa: F401
    except ImportError as e:
        raise ImportError(
            f"Cannot import sumolib from SUMO tools path: {tools_path}"
        ) from e

    import sumolib
    return sumolib


def get_neighboring_edges(net, x, y, radius):
    """
    Compatibility wrapper for different SUMO versions.
    """
    try:
        return net.getNeighboringEdges(x, y, radius, includeJunctions=False)
    except TypeError:
        return net.getNeighboringEdges(x, y, radius)


def lane_is_usable(lane):
    """
    Keep only lanes that are usable for road vehicles.
    """
    try:
        return not lane.allows("pedestrian")
    except Exception:
        return True


def parse_german_float(value):
    """
    Convert German decimal comma strings to float.
    Example: '11,28301180' -> 11.28301180
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip().replace(",", ".")

    try:
        return float(value)
    except ValueError:
        return None

def make_sensor_id(tk_blatt, bast_nr):
    """
    sensor_id = TK-Blatt + BASt-Nr.
    Example: 7734 + 9810 -> 77349810
    """
    return f"{int(tk_blatt)}{int(bast_nr):04d}"

def read_sensor_table(sensor_xlsx):
    if not os.path.exists(sensor_xlsx):
        print("[ERROR] Sensor file not found:")
        print(sensor_xlsx)
        return []

    workbook = load_workbook(sensor_xlsx, data_only=True)
    sheet = workbook.active

    headers = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        headers[str(cell.value).strip()] = col_idx

    required_cols = [
        "Richtung1_BASt-Nr.",
        "Richtung1_TK-Blatt",
        "Richtung1_Koor_WGS84_Lon",
        "Richtung1_Koor_WGS84_Lat",
        "Richtung2_BASt-Nr.",
        "Richtung2_TK-Blatt",
        "Richtung2_Koor_WGS84_Lon",
        "Richtung2_Koor_WGS84_Lat",
    ]

    for col in required_cols:
        if col not in headers:
            raise ValueError(f"Missing required column in Excel: {col}")

    sensors = []

    for row_idx in range(2, sheet.max_row + 1):
        r1_bast = sheet.cell(row_idx, headers["Richtung1_BASt-Nr."]).value
        r1_tk = sheet.cell(row_idx, headers["Richtung1_TK-Blatt"]).value

        r2_bast = sheet.cell(row_idx, headers["Richtung2_BASt-Nr."]).value
        r2_tk = sheet.cell(row_idx, headers["Richtung2_TK-Blatt"]).value

        lon1 = parse_german_float(sheet.cell(row_idx, headers["Richtung1_Koor_WGS84_Lon"]).value)
        lat1 = parse_german_float(sheet.cell(row_idx, headers["Richtung1_Koor_WGS84_Lat"]).value)

        lon2 = parse_german_float(sheet.cell(row_idx, headers["Richtung2_Koor_WGS84_Lon"]).value)
        lat2 = parse_german_float(sheet.cell(row_idx, headers["Richtung2_Koor_WGS84_Lat"]).value)

        if r1_bast is None or r1_tk is None:
            continue

        base_id_dir1 = make_sensor_id(r1_tk, r1_bast)

        if r2_bast is not None and r2_tk is not None:
            base_id_dir2 = make_sensor_id(r2_tk, r2_bast)
        else:
            base_id_dir2 = None

        if lon1 is not None and lat1 is not None:
            sensors.append({
                "sensor_id": f"{base_id_dir1}_dir1",
                "group_id": f"{base_id_dir1}_dir1",
                "direction_id": "dir1",
                "lon": lon1,
                "lat": lat1,
            })

        if base_id_dir2 is not None and lon2 is not None and lat2 is not None:
            sensors.append({
                "sensor_id": f"{base_id_dir2}_dir2",
                "group_id": f"{base_id_dir2}_dir2",
                "direction_id": "dir2",
                "lon": lon2,
                "lat": lat2,
            })

    return sensors


def lane_heading(lane):
    """
    Calculate lane direction angle in degrees.
    """
    shape = lane.getShape()
    if len(shape) < 2:
        return None

    x1, y1 = shape[0]
    x2, y2 = shape[-1]

    return math.degrees(math.atan2(y2 - y1, x2 - x1)) % 360


def angle_diff(a, b):
    """
    Smallest difference between two angles.
    """
    return abs((a - b + 180) % 360 - 180)


def get_lane_match(edge, lane, x, y, edge_dist, radius):
    lane_length = lane.getLength()

    try:
        lane_pos, lane_dist = lane.getClosestLanePosAndDist((x, y))
    except Exception:
        lane_pos = lane_length / 2.0
        lane_dist = edge_dist

    lane_pos = max(5.0, min(lane_pos, lane_length - 5.0))

    return {
        "edge_id": edge.getID(),
        "lane_id": lane.getID(),
        "lane_pos": round(lane_pos, 2),
        "distance_m": round(float(lane_dist), 2),
        "search_radius_m": radius,
        "heading": lane_heading(lane),
    }


def find_nearest_edge_for_sensor(net, lon, lat, search_radii=None):
    if search_radii is None:
        search_radii = [60, 100, 200, 400, 800, 1200]

    x, y = net.convertLonLat2XY(lon, lat)

    for radius in search_radii:
        candidates = []

        for edge, edge_dist in get_neighboring_edges(net, x, y, radius):
            if edge.isSpecial():
                continue

            usable_lanes = [
                lane for lane in edge.getLanes()
                if lane_is_usable(lane) and lane.getLength() >= 10
            ]

            if not usable_lanes:
                continue

            ref_lane = usable_lanes[0]

            try:
                _, lane_dist = ref_lane.getClosestLanePosAndDist((x, y))
            except Exception:
                lane_dist = edge_dist

            candidates.append((edge, float(lane_dist), radius))

        if candidates:
            edge, dist, radius = min(candidates, key=lambda item: item[1])
            return edge, dist, radius, x, y

    return None

def write_detector_additional_file(detector_records, additional_file, detector_output_file, freq=1800):
    """
    Write SUMO induction-loop detectors into an additional XML file.
    """
    root = ET.Element("additional")

    for record in detector_records:
        if not record["matched"]:
            continue

        ET.SubElement(
            root,
            "inductionLoop",
            {
                "id": record["sensor_id"],
                "lane": record["lane_id"],
                "pos": str(record["lane_pos"]),
                "freq": str(freq),
                "file": detector_output_file,
                "friendlyPos": "true",
            }
        )

    tree = ET.ElementTree(root)
    ET.indent(tree, space="    ")
    tree.write(additional_file, encoding="utf-8", xml_declaration=True)


def write_mapping_csv(detector_records, mapping_csv):
    """
    Export sensor-to-lane mapping results for manual checking.
    """
    fieldnames = [
        "detector_id",
        "group_id",
        "direction_id",
        "lon",
        "lat",
        "matched",
        "edge_id",
        "lane_id",
        "lane_pos",
        "distance_m",
        "search_radius_m",
    ]

    with open(mapping_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for record in detector_records:
            writer.writerow({
                "detector_id": record.get("detector_id", record["sensor_id"]),
                "group_id": record.get("group_id", ""),
                "direction_id": record.get("direction_id", ""),
                "lon": record["lon"],
                "lat": record["lat"],
                "matched": record["matched"],
                "edge_id": record.get("edge_id", ""),
                "lane_id": record.get("lane_id", ""),
                "lane_pos": record.get("lane_pos", ""),
                "distance_m": record.get("distance_m", ""),
                "search_radius_m": record.get("search_radius_m", ""),
            })

def aggregate_e1_output(detector_output_file, mapping_csv, aggregated_csv):
    detector_to_group = {}

    with open(mapping_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["matched"] == "True":
                detector_to_group[row["detector_id"]] = row["group_id"]

    tree = ET.parse(detector_output_file)
    root = tree.getroot()

    groups = {}

    for interval in root.findall("interval"):
        detector_id = interval.get("id")

        if detector_id not in detector_to_group:
            continue

        group_id = detector_to_group[detector_id]
        begin = interval.get("begin")
        end = interval.get("end")

        key = (begin, end, group_id)

        if key not in groups:
            groups[key] = {
                "begin": begin,
                "end": end,
                "group_id": group_id,
                "flow_sum": 0.0,
                "nVehContrib_sum": 0.0,
                "speed_weighted_sum": 0.0,
            }

        flow = float(interval.get("flow", 0))
        nveh = float(interval.get("nVehContrib", 0))
        speed = float(interval.get("speed", 0))

        groups[key]["flow_sum"] += flow
        groups[key]["nVehContrib_sum"] += nveh
        groups[key]["speed_weighted_sum"] += speed * nveh

    with open(aggregated_csv, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "begin",
            "end",
            "group_id",
            "flow_sum",
            "nVehContrib_sum",
            "speed_weighted_avg",
        ]

        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for _, value in sorted(groups.items()):
            if value["nVehContrib_sum"] > 0:
                speed_avg = value["speed_weighted_sum"] / value["nVehContrib_sum"]
            else:
                speed_avg = 0

            writer.writerow({
                "begin": value["begin"],
                "end": value["end"],
                "group_id": value["group_id"],
                "flow_sum": round(value["flow_sum"], 3),
                "nVehContrib_sum": round(value["nVehContrib_sum"], 3),
                "speed_weighted_avg": round(speed_avg, 3),
            })

def add_detectors():
    # =====================================
    # 1. Project paths
    # =====================================
    project_root = r"D:\SUMO_A9_Project"

    net_file = os.path.join(project_root, "sumo", "a8_corridor.net.xml")
    sensor_xlsx = os.path.join(project_root, "detectors", "5_detector_metering_points.xlsx")

    detector_dir = os.path.join(project_root, "detectors")
    os.makedirs(detector_dir, exist_ok=True)

    additional_file = os.path.join(detector_dir, "a8_detectors_secondary.add.xml")
    mapping_csv = os.path.join(detector_dir, "a8_detector_mapping_secondary.csv")
    detector_output_file = os.path.join(detector_dir, "a8_e1_output_secondary.xml")

    # =====================================
    # 2. Check files    # =====================================
    if not os.path.exists(net_file):
        print("[ERROR] Network file not found:")
        print(net_file)
        return

    if not os.path.exists(sensor_xlsx):
        print("[ERROR] Sensor Excel file not found:")
        print(sensor_xlsx)
        return

    # =====================================
    # 3. Load SUMO tools
    # =====================================
    sumolib = load_sumo_tools()

    print("=====================================")
    print("Loading SUMO network and sensor table")
    print(f"Network file : {net_file}")
    print(f"Sensor file  : {sensor_xlsx}")
    print("=====================================")

    net = sumolib.net.readNet(net_file)
    sensors = read_sensor_table(sensor_xlsx)

    print(f"[INFO] Sensors loaded: {len(sensors)}")

    if not sensors:
        print("[ERROR] No usable sensors found in the Excel file.")
        return

    # =====================================
    # 4. Match sensors to nearest lanes
    # =====================================
    detector_records = []

    for sensor in sensors:
        match = find_nearest_edge_for_sensor(
            net=net,
            lon=sensor["lon"],
            lat=sensor["lat"]
        )

        if match is None:
            record = dict(sensor)
            record["detector_id"] = f"{sensor['sensor_id']}_lane0"
            record["matched"] = False
            detector_records.append(record)

            print(f"[WARNING] No edge match found for sensor: {sensor['sensor_id']}")
            continue

        edge, edge_dist, radius, x, y = match

        lane_idx = 0

        for lane in edge.getLanes():
            if not lane_is_usable(lane):
                continue

            if lane.getLength() < 10:
                continue

            detector_id = f"{sensor['sensor_id']}_lane{lane_idx}"

            record = dict(sensor)
            record["sensor_id"] = detector_id
            record["detector_id"] = detector_id
            record["group_id"] = sensor["group_id"]
            record["direction_id"] = sensor["direction_id"]
            record["matched"] = True

            record.update(
                get_lane_match(
                    edge=edge,
                    lane=lane,
                    x=x,
                    y=y,
                    edge_dist=edge_dist,
                    radius=radius,
                )
            )

            detector_records.append(record)

            print(
                f"[MATCH] {detector_id} -> {record['lane_id']} "
                f"(group = {record['group_id']}, edge = {record['edge_id']})"
            )

            lane_idx += 1

    # =====================================
    # 5. Write outputs
    # =====================================
    write_detector_additional_file(
        detector_records=detector_records,
        additional_file=additional_file,
        detector_output_file=detector_output_file,
        freq=1800
    )

    write_mapping_csv(
        detector_records=detector_records,
        mapping_csv=mapping_csv
    )

    matched_count = sum(1 for r in detector_records if r["matched"])

    print("=====================================")
    print("[SUCCESS] Detector generation completed.")
    print(f"Matched lane detectors : {matched_count} / {len(detector_records)}")
    print(f"Detector file        : {additional_file}")
    print(f"Mapping CSV          : {mapping_csv}")
    print(f"Detector output file : {detector_output_file}")
    print("=====================================")


if __name__ == "__main__":
    add_detectors()