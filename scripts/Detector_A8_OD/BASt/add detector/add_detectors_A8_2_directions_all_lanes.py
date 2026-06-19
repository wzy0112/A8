import os
import sys
import csv
import math
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

# feq=1h
SUMO_HOME = r"D:\Eclipse\SUMO"

# 之前只有15个 而且是某一个lane上的 但是BASt网站里面每个detector给了2个richtunged 所以我必须给每个detector匹配上相邻反方向道路也放一个 一共30个
# 之前30个detecotrs只占据单独lane， 但是但是BASt网站里面每个detector给了2个richtungen, 对应的是整个edge的数据
# 9988发现其实能匹配相反方向 结果每匹配上 所以手动直接再addxml和mappingxsxl中指定edge -26652961 lane  -26652961_0 不管了 就这么继续干
def load_sumo_tools():
    tools_path = os.path.join(SUMO_HOME, "tools")
    if tools_path not in sys.path:
        sys.path.append(tools_path)

    try:
        import sumolib  # noqa: F401
    except ImportError as e:
        raise ImportError(
            f"Cannot import sumolib from SUMO tools path: {tools_path}"
        ) from e

    import sumolib
    return sumolib


def get_neighboring_edges(net, x, y, radius):
    """
    Compatibility wrapper for different SUMO versions.
    """
    try:
        return net.getNeighboringEdges(x, y, radius, includeJunctions=False)
    except TypeError:
        return net.getNeighboringEdges(x, y, radius)


def lane_is_usable(lane):
    """
    Keep only lanes that are usable for road vehicles.
    """
    try:
        return not lane.allows("pedestrian")
    except Exception:
        return True


def parse_german_float(value):
    """
    Convert German decimal comma strings to float.
    Example: '11,28301180' -> 11.28301180
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip().replace(",", ".")

    try:
        return float(value)
    except ValueError:
        return None


def read_sensor_table(sensor_xlsx):
    """
    Read new BASt detector Excel.

    Required columns:
    - BASt-Nr.
    - Koor_WGS84_Lon
    - Koor_WGS84_Lat
    """
    if not os.path.exists(sensor_xlsx):
        print("[ERROR] Sensor file not found:")
        print(sensor_xlsx)
        return []

    workbook = load_workbook(sensor_xlsx, data_only=True)
    sheet = workbook.active

    headers = {}
    for col_idx, cell in enumerate(sheet[1], start=1):
        headers[str(cell.value).strip()] = col_idx

    required_cols = ["BASt-Nr.", "Koor_WGS84_Lon", "Koor_WGS84_Lat"]

    for col in required_cols:
        if col not in headers:
            raise ValueError(f"Missing required column in Excel: {col}")

    sensors = []

    for row_idx in range(2, sheet.max_row + 1):
        sensor_id = sheet.cell(row_idx, headers["BASt-Nr."]).value
        lon_raw = sheet.cell(row_idx, headers["Koor_WGS84_Lon"]).value
        lat_raw = sheet.cell(row_idx, headers["Koor_WGS84_Lat"]).value

        lon = parse_german_float(lon_raw)
        lat = parse_german_float(lat_raw)

        if sensor_id is None or lon is None or lat is None:
            continue

        sensors.append({
            "sensor_id": str(sensor_id).strip(),
            "lon": lon,
            "lat": lat,
        })

    return sensors


def lane_heading(lane):
    """
    Calculate lane direction angle in degrees.
    """
    shape = lane.getShape()
    if len(shape) < 2:
        return None

    x1, y1 = shape[0]
    x2, y2 = shape[-1]

    return math.degrees(math.atan2(y2 - y1, x2 - x1)) % 360


def angle_diff(a, b):
    """
    Smallest difference between two angles.
    """
    return abs((a - b + 180) % 360 - 180)


def get_lane_match(edge, lane, x, y, edge_dist, radius):
    lane_length = lane.getLength()

    try:
        lane_pos, lane_dist = lane.getClosestLanePosAndDist((x, y))
    except Exception:
        lane_pos = lane_length / 2.0
        lane_dist = edge_dist

    lane_pos = max(5.0, min(lane_pos, lane_length - 5.0))

    return {
        "edge_id": edge.getID(),
        "lane_id": lane.getID(),
        "lane_pos": round(lane_pos, 2),
        "distance_m": round(float(lane_dist), 2),
        "search_radius_m": radius,
        "heading": lane_heading(lane),
    }


def find_two_direction_edges_for_sensor(net, lon, lat, search_radii=None):
    if search_radii is None:
        search_radii = [60, 100, 200, 400, 800, 1200]

    x, y = net.convertLonLat2XY(lon, lat)
    edge_candidates = []

    for radius in search_radii:
        neighboring_edges = get_neighboring_edges(net, x, y, radius)

        for edge, edge_dist in neighboring_edges:
            if edge.isSpecial():
                continue

            usable_lanes = [
                lane for lane in edge.getLanes()
                if lane_is_usable(lane) and lane.getLength() >= 10
            ]

            if not usable_lanes:
                continue

            ref_lane = usable_lanes[0]
            heading = lane_heading(ref_lane)

            if heading is None:
                continue

            try:
                _, lane_dist = ref_lane.getClosestLanePosAndDist((x, y))
            except Exception:
                lane_dist = edge_dist

            edge_candidates.append({
                "edge": edge,
                "edge_id": edge.getID(),
                "distance_m": float(lane_dist),
                "search_radius_m": radius,
                "heading": heading,
            })

        if len(edge_candidates) >= 2:
            break

    if not edge_candidates:
        return []

    main_edge = min(edge_candidates, key=lambda c: c["distance_m"])

    opposite_candidates = [
        c for c in edge_candidates
        if c["edge_id"] != main_edge["edge_id"]
        and angle_diff(c["heading"], main_edge["heading"]) >= 150
    ]

    selected_edges = [main_edge]

    if opposite_candidates:
        selected_edges.append(
            min(opposite_candidates, key=lambda c: c["distance_m"])
        )

    direction_results = []

    for dir_idx, edge_info in enumerate(selected_edges, start=1):
        edge = edge_info["edge"]
        lanes = []

        for lane in edge.getLanes():
            if not lane_is_usable(lane):
                continue

            if lane.getLength() < 10:
                continue

            lanes.append(
                get_lane_match(
                    edge=edge,
                    lane=lane,
                    x=x,
                    y=y,
                    edge_dist=edge_info["distance_m"],
                    radius=edge_info["search_radius_m"],
                )
            )

        direction_results.append({
            "direction_id": f"dir{dir_idx}",
            "group_id_suffix": f"dir{dir_idx}",
            "edge_id": edge.getID(),
            "heading": edge_info["heading"],
            "lanes": lanes,
        })

    return direction_results

def write_detector_additional_file(detector_records, additional_file, detector_output_file, freq=1800):
    """
    Write SUMO induction-loop detectors into an additional XML file.
    """
    root = ET.Element("additional")

    for record in detector_records:
        if not record["matched"]:
            continue

        ET.SubElement(
            root,
            "inductionLoop",
            {
                "id": record["sensor_id"],
                "lane": record["lane_id"],
                "pos": str(record["lane_pos"]),
                "freq": str(freq),
                "file": detector_output_file,
                "friendlyPos": "true",
            }
        )

    tree = ET.ElementTree(root)
    ET.indent(tree, space="    ")
    tree.write(additional_file, encoding="utf-8", xml_declaration=True)


def write_mapping_csv(detector_records, mapping_csv):
    """
    Export sensor-to-lane mapping results for manual checking.
    """
    fieldnames = [
        "detector_id",
        "group_id",
        "direction_id",
        "lon",
        "lat",
        "matched",
        "edge_id",
        "lane_id",
        "lane_pos",
        "distance_m",
        "search_radius_m",
    ]

    with open(mapping_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for record in detector_records:
            writer.writerow({
                "detector_id": record.get("detector_id", record["sensor_id"]),
                "group_id": record.get("group_id", ""),
                "direction_id": record.get("direction_id", ""),
                "lon": record["lon"],
                "lat": record["lat"],
                "matched": record["matched"],
                "edge_id": record.get("edge_id", ""),
                "lane_id": record.get("lane_id", ""),
                "lane_pos": record.get("lane_pos", ""),
                "distance_m": record.get("distance_m", ""),
                "search_radius_m": record.get("search_radius_m", ""),
            })

def aggregate_e1_output(detector_output_file, mapping_csv, aggregated_csv):
    detector_to_group = {}

    with open(mapping_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["matched"] == "True":
                detector_to_group[row["detector_id"]] = row["group_id"]

    tree = ET.parse(detector_output_file)
    root = tree.getroot()

    groups = {}

    for interval in root.findall("interval"):
        detector_id = interval.get("id")

        if detector_id not in detector_to_group:
            continue

        group_id = detector_to_group[detector_id]
        begin = interval.get("begin")
        end = interval.get("end")

        key = (begin, end, group_id)

        if key not in groups:
            groups[key] = {
                "begin": begin,
                "end": end,
                "group_id": group_id,
                "flow_sum": 0.0,
                "nVehContrib_sum": 0.0,
                "speed_weighted_sum": 0.0,
            }

        flow = float(interval.get("flow", 0))
        nveh = float(interval.get("nVehContrib", 0))
        speed = float(interval.get("speed", 0))

        groups[key]["flow_sum"] += flow
        groups[key]["nVehContrib_sum"] += nveh
        groups[key]["speed_weighted_sum"] += speed * nveh

    with open(aggregated_csv, "w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "begin",
            "end",
            "group_id",
            "flow_sum",
            "nVehContrib_sum",
            "speed_weighted_avg",
        ]

        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for _, value in sorted(groups.items()):
            if value["nVehContrib_sum"] > 0:
                speed_avg = value["speed_weighted_sum"] / value["nVehContrib_sum"]
            else:
                speed_avg = 0

            writer.writerow({
                "begin": value["begin"],
                "end": value["end"],
                "group_id": value["group_id"],
                "flow_sum": round(value["flow_sum"], 3),
                "nVehContrib_sum": round(value["nVehContrib_sum"], 3),
                "speed_weighted_avg": round(speed_avg, 3),
            })

def add_detectors():
    # =====================================
    # 1. Project paths
    # =====================================
    project_root = r"D:\SUMO_A9_Project"

    net_file = os.path.join(project_root, "sumo", "a8_corridor.net.xml")
    sensor_xlsx = os.path.join(project_root, "detectors", "15_detector_metering_points.xlsx")

    detector_dir = os.path.join(project_root, "detectors")
    os.makedirs(detector_dir, exist_ok=True)

    additional_file = os.path.join(detector_dir, "a8_detectors.add.xml")
    mapping_csv = os.path.join(detector_dir, "a8_detector_mapping.csv")
    detector_output_file = os.path.join(detector_dir, "a8_e1_output.xml")

    # =====================================
    # 2. Check files    # =====================================
    if not os.path.exists(net_file):
        print("[ERROR] Network file not found:")
        print(net_file)
        return

    if not os.path.exists(sensor_xlsx):
        print("[ERROR] Sensor Excel file not found:")
        print(sensor_xlsx)
        return

    # =====================================
    # 3. Load SUMO tools
    # =====================================
    sumolib = load_sumo_tools()

    print("=====================================")
    print("Loading SUMO network and sensor table")
    print(f"Network file : {net_file}")
    print(f"Sensor file  : {sensor_xlsx}")
    print("=====================================")

    net = sumolib.net.readNet(net_file)
    sensors = read_sensor_table(sensor_xlsx)

    print(f"[INFO] Sensors loaded: {len(sensors)}")

    if not sensors:
        print("[ERROR] No usable sensors found in the Excel file.")
        return

    # =====================================
    # 4. Match sensors to nearest lanes
    # =====================================
    detector_records = []

    for sensor in sensors:
        directions = find_two_direction_edges_for_sensor(
            net=net,
            lon=sensor["lon"],
            lat=sensor["lat"]
        )

        if not directions:
            record = dict(sensor)
            record["sensor_id"] = f"{sensor['sensor_id']}_dir1_lane0"
            record["detector_id"] = record["sensor_id"]
            record["group_id"] = f"{sensor['sensor_id']}_dir1"
            record["direction_id"] = "dir1"
            record["matched"] = False
            detector_records.append(record)

            print(f"[WARNING] No edge match found for sensor: {sensor['sensor_id']}")
            continue

        for direction in directions:
            direction_id = direction["direction_id"]
            group_id = f"{sensor['sensor_id']}_{direction_id}"

            for lane_idx, lane_match in enumerate(direction["lanes"]):
                record = dict(sensor)

                detector_id = f"{sensor['sensor_id']}_{direction_id}_lane{lane_idx}"

                record["sensor_id"] = detector_id
                record["detector_id"] = detector_id
                record["group_id"] = group_id
                record["direction_id"] = direction_id
                record["matched"] = True
                record.update(lane_match)

                detector_records.append(record)

                print(
                    f"[MATCH] {detector_id} -> {record['lane_id']} "
                    f"(group = {group_id}, edge = {record['edge_id']})"
                )

        if len(directions) == 1:
            print(f"[WARNING] Only one direction found for sensor: {sensor['sensor_id']}")


    # =====================================
    # 5. Write outputs
    # =====================================
    write_detector_additional_file(
        detector_records=detector_records,
        additional_file=additional_file,
        detector_output_file=detector_output_file,
        freq=1800
    )

    write_mapping_csv(
        detector_records=detector_records,
        mapping_csv=mapping_csv
    )

    matched_count = sum(1 for r in detector_records if r["matched"])

    print("=====================================")
    print("[SUCCESS] Detector generation completed.")
    pprint(f"Matched lane detectors : {matched_count} / {len(detector_records)}")
    print(f"Detector file        : {additional_file}")
    print(f"Mapping CSV          : {mapping_csv}")
    print(f"Detector output file : {detector_output_file}")
    print("=====================================")


if __name__ == "__main__":
    add_detectors()